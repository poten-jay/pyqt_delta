
import json
from functools import partial
import openpyxl
from PyQt5.QtWidgets import  QPushButton, QMainWindow, QLabel, QDialog, QDoubleSpinBox
from PyQt5.QtCore import  Qt, QTimer, pyqtSignal
from PyQt5.QtGui import QPixmap
from datetime import datetime

# from publisher import GUI_Node
from manual import MyApp
from home import MyHome
from move import MyMove
from info import MyInfo

import setting
from trajectory import *


excel_file_path = "cal.xlsx" 


wb = openpyxl.load_workbook(excel_file_path)
sheet = wb["angles"]


q_values = []  # IKinem 함수의 결과를 저장할 리스트


for i in range(1,26):
    value1 = sheet.cell(row=i + 1, column=1).value
    value2 = sheet.cell(row=i + 1, column=2).value
    value3 = sheet.cell(row=i + 1, column=3).value

    q_value = [IKinem(value1,value2,value3)]  # 예시로 임시 값을 할당
    q_values.append(q_value)  # 결과를 리스트에 추가

    # print(f"Value1: {value1}")
    # print(f"Value2: {value2}")
    # print(f"Value3: {value3}")

# q_values 리스트에 저장된 IKinem 함수의 결과 출력
for q_value in q_values:
    print(f"IKinem Result: {q_value}")


    # q_value = list(IKinem(value1,value2,value3))
    # print(q_value)

print('0000000' , q_values[0][0][0])
print('11111111' , q_values[0][0][1])
print('22222222' , q_values[0][0][2])

print('0000000' , q_values[24][0][0])
print('1111111' , q_values[24][0][1])
print('2222222' , q_values[24][0][2])



