
import json
from functools import partial
import openpyxl
from PyQt5.QtWidgets import  QPushButton, QMainWindow, QLabel, QDialog, QDoubleSpinBox
from PyQt5.QtCore import  Qt, QTimer, pyqtSignal
from PyQt5.QtGui import QPixmap
from datetime import datetime

# from publisher import GUI_Node
from manual import MyApp
from home import MyHome
from move import MyMove
from info import MyInfo

import setting
from trajectory import *


class MyCal(QMainWindow):
    goToStartScreen = pyqtSignal()
    def __init__(self,node, parent=None):
        super().__init__()
        self.node = node
        self.parent = parent
        self.initUI()
        
    def initUI(self):

        # Excel 파일 경로 및 파일 이름 지정
        excel_file_path = "document/cal.xlsx"
        # 만약 Excel 파일이 없으면 기본 포멧으로 만들기
        if not os.path.exists(excel_file_path):
            self.create_initial_excel(excel_file_path)

        # UI 초기화 코드
            
###### 배경화면 지정 ##################################################################
        original_pixmap = QPixmap("img/cal.png")
        scaled_pixmap = original_pixmap.scaled(800, 600, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        # QLabel 생성 및 QPixmap 설정
        lbl_img = QLabel(self)
        lbl_img.setPixmap(scaled_pixmap)
        lbl_img.setGeometry(0, 0, scaled_pixmap.width(), scaled_pixmap.height())

####### 버튼 좌표 설정 기준 ###########################################################
        btn_x = 100
        btn_y = 50
####### bnt_1 #####################################################################
        self.cal_1 = QPushButton('01', self)
        self.cal_1.clicked.connect(partial(self.readexcel, 2))
        self.cal_1.setGeometry(btn_x+200, btn_y+100, 50, 50)

####### btn_2 #####################################################################
        self.cal_2 = QPushButton('02', self)
        self.cal_2.clicked.connect(partial(self.readexcel, 3))
        self.cal_2.setGeometry(btn_x+275, btn_y+100, 50, 50)

####### btn_3 #####################################################################
        self.cal_3 = QPushButton('03', self)
        self.cal_3.clicked.connect(partial(self.readexcel, 4))
        self.cal_3.setGeometry(btn_x+350, btn_y+100, 50, 50)

####### btn_4 #####################################################################
        self.cal_4 = QPushButton('04', self)
        self.cal_4.clicked.connect(partial(self.readexcel, 5))
        self.cal_4.setGeometry(btn_x+425, btn_y+100, 50, 50)

####### btn_5 #####################################################################
        self.cal_5 = QPushButton('05', self)
        self.cal_5.clicked.connect(partial(self.readexcel, 6))
        self.cal_5.setGeometry(btn_x+500, btn_y+100, 50, 50)

####### bnt_6 #####################################################################
        self.cal_6 = QPushButton('06', self)
        self.cal_6.clicked.connect(partial(self.readexcel, 7))
        self.cal_6.setGeometry(btn_x+200, btn_y+175, 50, 50)

####### btn_7 #####################################################################
        self.cal_7 = QPushButton('07', self)
        self.cal_7.clicked.connect(partial(self.readexcel, 8))
        self.cal_7.setGeometry(btn_x+275, btn_y+175, 50, 50)

####### btn_8 ####################################################################
        self.cal_8 = QPushButton('08', self)
        self.cal_8.clicked.connect(partial(self.readexcel, 9))
        self.cal_8.setGeometry(btn_x+350, btn_y+175, 50, 50)

####### btn_9 #####################################################################
        self.cal_9 = QPushButton('09', self)
        self.cal_9.clicked.connect(partial(self.readexcel, 10))
        self.cal_9.setGeometry(btn_x+425, btn_y+175, 50, 50)

####### btn_10 #####################################################################
        self.cal_10 = QPushButton('10', self)
        self.cal_10.clicked.connect(partial(self.readexcel, 11))
        self.cal_10.setGeometry(btn_x+500, btn_y+175, 50, 50)

####### 나머지 btn ####################################################################
        self.cal_11 = QPushButton('11', self)
        self.cal_11.clicked.connect(partial(self.readexcel, 12))
        self.cal_11.setGeometry(btn_x+200, btn_y+250, 50, 50)

        self.cal_12 = QPushButton('12', self)
        self.cal_12.clicked.connect(partial(self.readexcel, 13))
        self.cal_12.setGeometry(btn_x+275, btn_y+250, 50, 50)

        self.cal_13 = QPushButton('13', self)
        self.cal_13.clicked.connect(partial(self.readexcel, 14))
        self.cal_13.setGeometry(btn_x+350, btn_y+250, 50, 50)

        self.cal_14 = QPushButton('14', self)
        self.cal_14.clicked.connect(partial(self.readexcel, 15))
        self.cal_14.setGeometry(btn_x+425, btn_y+250, 50, 50)

        self.cal_15 = QPushButton('15', self)
        self.cal_15.clicked.connect(partial(self.readexcel, 16))
        self.cal_15.setGeometry(btn_x+500, btn_y+250, 50, 50)

        self.cal_16 = QPushButton('16', self)
        self.cal_16.clicked.connect(partial(self.readexcel, 17))
        self.cal_16.setGeometry(btn_x+200, btn_y+325, 50, 50)

        self.cal_17 = QPushButton('17', self)
        self.cal_17.clicked.connect(partial(self.readexcel, 18))
        self.cal_17.setGeometry(btn_x+275, btn_y+325, 50, 50)

        self.cal_18 = QPushButton('18', self)
        self.cal_18.clicked.connect(partial(self.readexcel, 19))
        self.cal_18.setGeometry(btn_x+350, btn_y+325, 50, 50)

        self.cal_19 = QPushButton('19', self)
        self.cal_19.clicked.connect(partial(self.readexcel, 20))
        self.cal_19.setGeometry(btn_x+425, btn_y+325, 50, 50)

        self.cal_20 = QPushButton('20', self)
        self.cal_20.clicked.connect(partial(self.readexcel, 21))
        self.cal_20.setGeometry(btn_x+500, btn_y+325, 50, 50)

        self.cal_21 = QPushButton('21', self)
        self.cal_21.clicked.connect(partial(self.readexcel, 22))
        self.cal_21.setGeometry(btn_x+200, btn_y+400, 50, 50)

        self.cal_22 = QPushButton('22', self)
        self.cal_22.clicked.connect(partial(self.readexcel, 23))
        self.cal_22.setGeometry(btn_x+275, btn_y+400, 50, 50)

        self.cal_23 = QPushButton('23', self)
        self.cal_23.clicked.connect(partial(self.readexcel, 24))
        self.cal_23.setGeometry(btn_x+350, btn_y+400, 50, 50)

        self.cal_24 = QPushButton('24', self)
        self.cal_24.clicked.connect(partial(self.readexcel, 25))
        self.cal_24.setGeometry(btn_x+425, btn_y+400, 50, 50)

        self.cal_25 = QPushButton('25', self)
        self.cal_25.clicked.connect(partial(self.readexcel, 26))
        self.cal_25.setGeometry(btn_x+500, btn_y+400, 50, 50)


####### 뒤로 가기 버튼 ###############################################################
        self.btnback = QPushButton('<<', self)
        self.btnback.clicked.connect(self.goToStartScreen.emit)
        
        self.btnback.setGeometry(0, 528, 50, 50)

####### 시간 #######################################################################
        # 시간 표시 라벨 설정
        self.time_label = QLabel(self)
        self.time_label.setGeometry(660, 585, 200, 10)
        self.time_label.setStyleSheet("font-size: 14px;")
        self.time_label.setStyleSheet("Color : white")

        # QTimer 설정
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_time)
        self.timer.start(1000)


####### 이미지 삽입 #################################################################
        
    def input_image(self, x, y, w, h, img_path):
        pixmap = QPixmap(img_path)
        scaled_pixmap = pixmap.scaled(w, h, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        # QLabel 생성 및 QPixmap 설정
        lbl_img = QLabel(self)
        lbl_img.setPixmap(scaled_pixmap)
        lbl_img.setGeometry(x, y, scaled_pixmap.width(), scaled_pixmap.height())

####### 시간 #################################################################
    def update_time(self):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.time_label.setText(current_time)

###### 엑셀읽어오기 #############################################################
    def readexcel(self, n):
        self.n = n
        
        # self.node.publish_xyz( x, y, z)  # Create an instance of xyz_button
        # Excel 파일에서 값을 읽어와서 표시하는 팝업 생성 및 표시
        excel_file_path = "document/cal.xlsx"  # 엑셀 파일 경로를 지정하세요.
        self.trajectory_1(excel_file_path)
        popup = ExcelPopup(excel_file_path, n, self.node, self)
        popup.exec_()


###### 기본 좌표 25에 대한 키네마틱스 - 세타값 ######################################
    def trajectory_1(self, excel_file_path):

        wb = openpyxl.load_workbook(excel_file_path)
        sheet = wb["angles"]

        q_values = []  # IKinem 함수의 결과를 저장할 리스트
        
        for i in range(1,26):
            value1 = sheet.cell(row=i + 1, column=1).value
            value2 = sheet.cell(row=i + 1, column=2).value
            value3 = sheet.cell(row=i + 1, column=3).value

            q_value = [IKinem(value1,value2,value3)]  # 예시로 임시 값을 할당
            q_values.append(q_value)  # 결과를 리스트에 추가

            

            for q_value in q_values:
                
                # IK_value_1 = "{:.15}".format(q_values[i-1][0][0])
                # IK_value_2 = "{:.15}".format(q_values[i-1][0][1])
                # IK_value_3 = "{:.15}".format(q_values[i-1][0][2])
                IK_value_1 = q_values[i-1][0][0]
                IK_value_2 = q_values[i-1][0][1]
                IK_value_3 = q_values[i-1][0][2]

                # 계산된 값을 D, E, F 열에 입력
                sheet.cell(row=i + 1, column=4, value=IK_value_1)
                sheet.cell(row=i + 1, column=5, value=IK_value_2)
                sheet.cell(row=i + 1, column=6, value=IK_value_3)

        # 변경된 내용을 엑셀 파일에 저장
        wb.save(excel_file_path)

        # 저장 후에는 엑셀 파일을 닫습니다.
        wb.close()


####### 엑셀파일이 없을때 기본 포멧 만들기 #############################################
    def create_initial_excel(self, excel_file_path):
        # Create a new Excel file and add initial values
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "angles"

        # 초기값 설정
        x = -400
        y = -400
        z = -900

        # 10x10 행렬 생성 및 Excel에 저장
        for i in range(25):
                # 현재 값을 시트에 추가
                sheet.cell(row=i + 2, column=1, value=x)
                sheet.cell(row=i + 2, column=2, value=y)
                sheet.cell(row=i + 2, column=3, value=z)

                sheet.cell(row=i + 2, column=7, value=x)
                sheet.cell(row=i + 2, column=8, value=y)
                sheet.cell(row=i + 2, column=9, value=z)

                # y 값을 증가
                y += 200

                # x 값을 증가, y는 초기값으로 재설정
                if (i + 1) % 5 == 0:
                        x += 200
                        y = -400

        # Excel 파일 저장
        wb.save(excel_file_path)
        wb.close()



####### 나머지 코드는 그대로 두고 아래의 ExcelPopup 클래스 추가 ##########################


# x = 0.0
# y = 0.0
# z = -900.0

###### 팝업창 설정 ################################################################
class ExcelPopup(QDialog):
    # btn = xyz_button(x, y, z)

    def __init__(self, excel_file, row_index, node, parent=None):
        super().__init__(parent)
        self.excel_file = excel_file
        self.row_index = row_index
        self.node = node
        # self.xyz_btn = xyz_btn_instance  # Store the xyz_button instance
        self.initUI()
        self.load_excel_data()

    def initUI(self):
        self.setWindowTitle("CALIBRATION")
        self.setGeometry(100, 100, 400, 200)

        # Update Button
        self.btnUpdate = QPushButton('MOVE', self)
        self.btnUpdate.setGeometry(100, 100, 50, 30)  # Adjust as needed
        # self.btnUpdate.clicked.connect(self.updateXYZ)
        # self.btnUpdate.clicked.connec:t(lambda: print('gogogoogo xyz'))
        self.btnUpdate.clicked.connect(self.on_run_button_clicked)  # Connect the button click signal

        # Update Button
        self.btnSave = QPushButton('SAVE', self)
        self.btnSave.setGeometry(200, 100, 50, 30)  # Adjust as needed
        # self.btnSave.clicked.connect(self.updateXYZ)
        self.btnSave.clicked.connect(self.save_spin_xyz_kinematics)

        # 더블 스핀 박스 추가
        self.double_spin_box1 = QDoubleSpinBox(self)
        self.double_spin_box2 = QDoubleSpinBox(self)
        self.double_spin_box3 = QDoubleSpinBox(self)
        self.double_spin_box1.setRange(setting.x_min, setting.x_max)
        self.double_spin_box2.setRange(setting.y_min, setting.y_max)
        self.double_spin_box3.setRange(setting.z_min, setting.z_max)
        # self.double_spin_box1.setRange(setting.x_min, setting.x_max)
        # self.double_spin_box1.setRange(setting.y_min, setting.y_max)
        # self.double_spin_box1.setRange(setting.z_min, setting.z_max)
        
        self.double_spin_box1.setGeometry(25, 50, 100, 30)
        self.double_spin_box2.setGeometry(150, 50, 100, 30)
        self.double_spin_box3.setGeometry(275, 50, 100, 30)
        
        self.double_spin_box1.setKeyboardTracking(False)
        self.double_spin_box2.setKeyboardTracking(False)
        self.double_spin_box3.setKeyboardTracking(False)

        wb = openpyxl.load_workbook(self.excel_file)
        sheet = wb["angles"]
        value1 = float(sheet[self.row_index][6].value)
        value2 = float(sheet[self.row_index][7].value)
        value3 = float(sheet[self.row_index][8].value)
        # print(value1,value2,value3)
        self.double_spin_box1.setValue(value1)
        self.double_spin_box2.setValue(value2)
        self.double_spin_box3.setValue(value3)

        # 스핀 박스 값 변경 시그널에 연결
        self.double_spin_box1.valueChanged.connect(self.on_spin_box_changed)
        self.double_spin_box2.valueChanged.connect(self.on_spin_box_changed)
        self.double_spin_box3.valueChanged.connect(self.on_spin_box_changed)

        # # 스핀 박스 값 변경 시그널에 연결
        # self.double_spin_box1.valueChanged.connect(self.spin_box_1)
        # self.double_spin_box2.valueChanged.connect(self.spin_box_2)
        # self.double_spin_box3.valueChanged.connect(self.spin_box_3)


    ### RUN 버튼을 클릭 했을때 #######################################################
    def on_run_button_clicked(self, value):
        print(f"Current value: {value}")
        # This method will be executed when the "Run" button is clicked
        x = self.double_spin_box1.value()
        y = self.double_spin_box2.value()
        z = self.double_spin_box3.value()
        self.node.publish_xyz(x,y,z)


    def load_excel_data(self):
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            sheet = wb["angles"]
            value1 = float(sheet[self.row_index][6].value)
            value2 = float(sheet[self.row_index][7].value)
            value3 = float(sheet[self.row_index][8].value)


            # 엑셀에서 읽어와서 스핀박스에 띄워줌
            self.double_spin_box1.setValue(value1)
                                    # Current value: -400.0
                                    # xㅌㅌ :  -400.0
                                    # yㅛㅛ :  0.0
                                    # zㅋㅋ :  -700.0
            self.double_spin_box2.setValue(value2)
                                    # Current value: -400.0
                                    # xㅌㅌ :  -400.0
                                    # yㅛㅛ :  -400.0
                                    # zㅋㅋ :  -700.0
            self.double_spin_box3.setValue(value3)
                                    # Current value: -900.0
                                    # xㅌㅌ :  -400.0
                                    # yㅛㅛ :  -400.0
                                    # zㅋㅋ :  -900.0


        except Exception as e:
            self.double_spin_box1.setValue(0.0)
            self.double_spin_box2.setValue(0.0)
            self.double_spin_box3.setValue(0.0)

    def on_spin_box_changed(self, value):
        # value 인수로 현재 스핀 박스의 값이 전달됩니다.
        print(f"Current value: {value}")
        # Update the xyz_button instance with new values
        # print(value)
        # if isinstance(value, float) ==True:
        x = self.double_spin_box1.value()
        print('xㅌㅌ : ', x)
        y = self.double_spin_box2.value()
        print('yㅛㅛ : ', y)
        z = self.double_spin_box3.value()
        print('zㅋㅋ : ',  z)
        self.node.publish_xyz(x,y,z)


    # def spin_box_1(self):
    #     x = self.double_spin_box1.value()
    #     self.node.publish_x(x)

    # def spin_box_2(self):
    #     y = self.double_spin_box2.value()
    #     self.node.publish_y(y)

    # def spin_box_3(self):
    #     z = self.double_spin_box3.value()
    #     self.node.publish_z(z)

    def save_spin_xyz_kinematics(self):
        excel_file_path = "document/cal.xlsx"
        x = self.double_spin_box1.value()
        y = self.double_spin_box2.value()
        z = self.double_spin_box3.value()

        wb = openpyxl.load_workbook(excel_file_path)
        sheet = wb["angles"]

        sheet.cell(row=self.row_index, column=7, value=x)
        sheet.cell(row=self.row_index, column=8, value=y)
        sheet.cell(row=self.row_index, column=9, value=z)

        # value1 = float(sheet[self.row_index][6].value)
        # value2 = float(sheet[self.row_index][7].value)
        # value3 = float(sheet[self.row_index][8].value)
       
        q_value = [IKinem(x,y,z)]
        print('Save : (row :',self.row_index,"),", q_value[0])

        sheet.cell(row=self.row_index, column=10, value=q_value[0][0])
        sheet.cell(row=self.row_index, column=11, value=q_value[0][1])
        sheet.cell(row=self.row_index, column=12, value=q_value[0][2])

        # 변경된 내용을 엑셀 파일에 저장
        wb.save(excel_file_path)

        # 저장 후에는 엑셀 파일을 닫습니다.
        wb.close()


##########################################################


    def onBackButtonClick(self):
        # Emit the signal when the button is clicked
        self.goToStartScreen.emit()


