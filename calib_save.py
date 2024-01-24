import rclpy
from rclpy.node import Node
from rclpy.duration import Duration
import time
import math
from trajectory import *
import argparse  

from sensor_msgs.msg import JointState 
from std_msgs.msg import Float64MultiArray

import threading

import os 
import pandas
import openpyxl
import sys

'''
# WIM Lab workspace
x= -460 ~ +460
y= -460 ~ +460
z= -1090 ~ -700
'''



MOVE_VEL = 300  # mm/s

class DeltaMover(Node):
    def __init__(self, args):
        super().__init__('move_w_delta')
        self.args = args

        self.publisher = self.create_publisher(Float64MultiArray, "joint_group_position_controller/commands", 10)
        self.subscription = self.create_subscription(JointState, '/joint_states', self.joint_state_callback, 10)
        # self.subscription  # prevent unused variable warning 
        # Initialize the parameter with an empty list  

        self.stop_program = False

                
        
        self.current_joint_position = []

        self.start_pos = (0, 0, -320)
        self.end_pos = (0, 0, -350)
        self.movement_done = True
        self.last_index = False

        self.keep_running = True
        self.lock = threading.Lock()

        self.movement_thread = threading.Thread(target=self.movement_loop)
        self.movement_thread.start()

                                                
        self.alternating_positions = []
        self.ordered_positions = []

        # self.fisrt_time = True
        self.t0 = 0
       
        self.alternating_index = 0
        self.start_time = self.get_clock().now()

        self.X = None
        self.Y = None
        self.Z = None


    

    def joint_state_callback(self, msg):
        with self.lock:
            if self.movement_done:
                self.current_joint_position = msg.position
                desired_order = ['joint1', 'joint2', 'joint3']
                self.ordered_positions = [0] * len(desired_order)

                # Reorder the joint states
                for i, name in enumerate(desired_order):
                    if name in msg.name:
                        index = msg.name.index(name)
                        self.ordered_positions[i] = msg.position[index]
     
                self.current_joint_position = [i for i in self.ordered_positions]

                mode = self.args.mode  # Use the 'mode' argument value
                self.X = self.args.x
                self.Y = self.args.y
                self.Z = self.args.z    

                print("X,Y,Z", self.X, self.Y, self.Z)

                if -460 > self.X or +460 < self.X:
                    raise Exception(" X range limit")
                elif -460 > self.Y or +460 < self.Y:
                    raise Exception(" Y range limit")
                elif -1090 > self.Z or -700 < self.Z:
                    raise Exception(" Z range limit") 
                

                # Check !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!


                


                q_value = list(IKinem(self.X,self.Y,self.Z))


                

                if mode == 0:
                    self.write_to_xlsx(self.X, self.Y, self.Z, q_value)
                elif mode == 1:
                    self.read_and_write_to_xlsx(self.X, self.Y, self.Z, q_value)


                start_pos = (0, 0, -300)
                end_pos = (0, 0, -500)        
                
                result0 = generate_trajectory(self.current_joint_position,
                                                            IKinem(self.X, self.Y, self.Z),
                                                            3,
                                                            0.001) # 0.001 
                
                

                self.alternating_positions.extend(result0)


                self.movement_done = False
                # self.t0 = self.get_clock().now()
    
    def write_to_xlsx(self, X, Y, Z, q_value):
        try:
            write_wb = openpyxl.load_workbook('output1.xlsx', data_only=True)
            if 'angles' in write_wb.sheetnames:
                write_ws = write_wb['angles']
            else:
                write_ws = write_wb.create_sheet('angles')
        except FileNotFoundError:
            write_wb = openpyxl.Workbook()
            write_ws = write_wb.create_sheet('angles')
        last_row = write_ws.max_row 
        empty_row = last_row + 1
        last_XYZ = [write_ws.cell(row=last_row, column=i).value for i in range(1, 4)]

        if last_XYZ != [X, Y, Z]:
            for i, data in enumerate([X, Y, Z] + q_value):
                cell = write_ws.cell(row=empty_row, column=i + 1)
                cell.value = data
            write_wb.save('output1.xlsx')
        else:
            self.get_logger().error('Error: XYZ values same with the last row in the file.')

    def read_and_write_to_xlsx(self, X, Y, Z, q_value):
        try:
            write_wb = openpyxl.load_workbook('output1.xlsx', data_only=True)
            if 'angles' in write_wb.sheetnames:
                write_ws = write_wb['angles']
            else:
                raise FileNotFoundError
        except FileNotFoundError:
            self.get_logger().error('File not found.')
            return

        last_row = write_ws.max_row
        last_XYZ = [write_ws.cell(row=last_row, column=i).value for i in range(1, 4)]

        if last_XYZ != [X, Y, Z]:
            for i, data in enumerate(q_value):
                cell = write_ws.cell(row=last_row, column=i + 7)
                cell.value = data
            write_wb.save('output1.xlsx')
        else:
            self.get_logger().error('Error: XYZ values same with the last row in the file. Retry !!')


    def movement_loop(self):
        while self.keep_running:
            
            print("continue")
            # time.sleep(0.1)
            with self.lock:
                if not self.alternating_positions or self.movement_done:
                # if self.movement_done:
                    continue
                
                if self.last_index:
                    self.publish_point()
                    self.last_index = False
                    
                    self.alternating_positions = []
                    
                    # self.movement_done = True
                    self.keep_running = False

                    self.stop_program = True


                    
                    continue

                

                self.publish_point()

                self.alternating_index = (self.alternating_index + 1) # % len(self.alternating_positions)
        
                if self.alternating_index == len(self.alternating_positions) - 1:
                    self.last_index = True
            
            
            time.sleep(0.0005) # Adjust the sleep time as needed. # 200hz=0.005


    

    def publish_point(self):

        positions = self.alternating_positions[self.alternating_index]
        msg = Float64MultiArray()
        msg.data = positions



        self.publisher.publish(msg)
        self.start_time = self.get_clock().now()


    def destroy_node(self):
        self.keep_running = False
        self.movement_thread.join()
        super().destroy_node()


def parse_arguments():
        # Create an argument parser
        parser = argparse.ArgumentParser(description='Move_w_delta program with mode option')

        # Add the 'mode' argument
        parser.add_argument('--mode', type=int, default=0, help='Set the mode (0 or 1)')
        parser.add_argument('--x', type=float, default=0, help='Set the x -200 ~ +200')
        parser.add_argument('--y', type=float, default=0, help='Set the y -200 ~ +200')
        parser.add_argument('--z', type=float, default=-900, help='Set the y -200 ~ +200')

        # Parse the command-line arguments
        args = parser.parse_args()

        return args
        


def main(args=None):
    pargs = parse_arguments()
    rclpy.init(args=args)
    delta_mover = DeltaMover(pargs)

    while rclpy.ok() and not delta_mover.stop_program:
        rclpy.spin_once(delta_mover, timeout_sec=0.1)

    # rclpy.spin(delta_mover)
    delta_mover.destroy_node()
    rclpy.shutdown()


if __name__ == '__main__':
    main()
